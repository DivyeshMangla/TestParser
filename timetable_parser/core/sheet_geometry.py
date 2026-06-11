from __future__ import annotations

from datetime import datetime, timedelta

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from timetable_parser.core.models import CellBounds, RawCell
from timetable_parser.extractors.day_slots import Slot


def build_merge_bounds(sheet: Worksheet) -> dict[tuple[int, int], CellBounds]:
    merge_bounds: dict[tuple[int, int], CellBounds] = {}
    for merged_range in sheet.merged_cells.ranges:
        bounds = CellBounds(
            min_row=merged_range.min_row,
            min_col=merged_range.min_col,
            max_row=merged_range.max_row,
            max_col=merged_range.max_col,
        )
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_bounds[(row, col)] = bounds
    return merge_bounds


def visible_bounds_for_cell(merge_bounds: dict[tuple[int, int], CellBounds], row: int, col: int) -> CellBounds:
    return merge_bounds.get((row, col), CellBounds(min_row=row, min_col=col, max_row=row, max_col=col))


def raw_cells_in_bounds(sheet: Worksheet, min_row: int, max_row: int, bounds: CellBounds) -> list[RawCell]:
    cells: list[RawCell] = []
    seen_coordinates: set[str] = set()

    for row in range(min_row, max_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            value = clean_text(sheet.cell(row=row, column=col).value)
            if value is None:
                continue

            coordinate = f"{get_column_letter(col)}{row}"
            if coordinate in seen_coordinates:
                continue

            seen_coordinates.add(coordinate)
            cells.append(RawCell(coordinate=coordinate, row=row, column=col, value=value))

    return cells


def rectangle_has_raw(sheet: Worksheet, min_row: int, max_row: int, bounds: CellBounds) -> bool:
    return bool(raw_cells_in_bounds(sheet, min_row, max_row, bounds))


def row_has_top_border(sheet: Worksheet, row: int, bounds: CellBounds) -> bool:
    for col in range(bounds.min_col, bounds.max_col + 1):
        border = sheet.cell(row=row, column=col).border.top
        if border and border.style:
            return True
    return False


def row_has_bottom_border(sheet: Worksheet, row: int, bounds: CellBounds) -> bool:
    for col in range(bounds.min_col, bounds.max_col + 1):
        border = sheet.cell(row=row, column=col).border.bottom
        if border and border.style:
            return True
    return False


def slot_end_row(slot: Slot, remaining_slots: list[Slot], offset: int) -> int:
    if offset + 1 < len(remaining_slots):
        return remaining_slots[offset + 1].cell.row - 1
    return slot.cell.row + 1


def dedupe_raw_cells(cells: list[RawCell]) -> list[RawCell]:
    seen: set[str] = set()
    unique: list[RawCell] = []
    for cell in sorted(cells, key=lambda item: (item.row, item.column)):
        if cell.coordinate in seen:
            continue
        seen.add(cell.coordinate)
        unique.append(cell)
    return unique


def dedupe_values(values: object) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        if not isinstance(value, str) or value in seen:
            continue
        seen.add(value)
        unique.append(value)
    return unique


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def end_time(start_time: str, periods: int) -> str:
    parsed = datetime.strptime(start_time, "%H:%M")
    return (parsed + timedelta(minutes=50 * periods)).strftime("%H:%M")
